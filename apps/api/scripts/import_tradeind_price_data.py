import html
import http.client
import json
import re
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from app.db import SessionLocal
from app.models import CanonicalProduct, ProductFamily, ProductPriceObservation, Retailer, Store
from app.package_parser import parse_package
from app.product_identity import build_product_identity, clean_product_name

SEARCH_URL = "https://tradeind.gov.tt/wp-json/wp/v2/search"
POST_URL = "https://tradeind.gov.tt/wp-json/wp/v2/posts/{post_id}"
RAW_DIR = Path(__file__).resolve().parents[3] / "data" / "tradeind" / "raw"
USER_AGENT = "groceries-mvp-tradeind-import/0.1"
REGION_SHEETS = {"west", "north", "east", "central", "south", "tobago"}

RETAILER_ALIASES = {
    "jta": "JTA Supermarkets",
    "jta supermarket": "JTA Supermarkets",
    "jta supermarkets": "JTA Supermarkets",
    "massy": "Massy Stores",
    "massy stores": "Massy Stores",
    "tru valu": "Tru Valu",
    "tru-valu": "Tru Valu",
    "xtra foods": "Xtra Foods",
    "persad d food king": "Persad's D' Food King",
    "persad's": "Persad's D' Food King",
    "persad's d food king": "Persad's D' Food King",
    "pennysavers": "Penny Savers",
    "penny savers": "Penny Savers",
    "cost cutters": "Cost Cutters",
    "coss cutters": "Cost Cutters",
    "food basket": "Food Basket",
}


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def request_json(url: str) -> Any:
    last_error: Exception | None = None
    for _ in range(3):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(req, timeout=90) as response:
                return json.loads(response.read().decode("utf-8"))
        except (TimeoutError, http.client.IncompleteRead, urllib.error.URLError) as exc:
            last_error = exc
    if last_error:
        raise last_error
    raise RuntimeError(f"Unable to request JSON from {url}")


def normalize(value: str | None) -> str:
    if not value:
        return ""
    value = html.unescape(str(value)).lower().replace("’", "'")
    value = re.sub(r"[^a-z0-9']+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def clean_download_url(value: str) -> str:
    value = html.unescape(urllib.parse.unquote(value)).strip().strip("”’'\" ")
    return value.replace("http://tradeind.gov.tt", "https://tradeind.gov.tt")


def extract_download_links(content: str) -> list[str]:
    content = html.unescape(content)
    found: list[str] = []
    found += re.findall(r"https?://[^\s\"'<>\]”]+?\.(?:xlsx|xls|pdf)", content, flags=re.I)
    for match in re.findall(r"url:([^\]\|\s\"']+)", content, flags=re.I):
        found.append(match)
    for href in re.findall(r"href=[\"']([^\"']+)[\"']", content, flags=re.I):
        found.append(href)
    cleaned = []
    for url in found:
        url = clean_download_url(url)
        if re.search(r"\.(xlsx|xls|pdf)(\?|$)", url, re.I):
            cleaned.append(url)
    return sorted(set(cleaned))


def discover_price_files() -> list[dict[str, Any]]:
    posts = []
    for page in range(1, 10):
        url = SEARCH_URL + "?" + urllib.parse.urlencode({"search": "Supermarket Prices", "per_page": 100, "page": page})
        data = request_json(url)
        if not data:
            break
        posts.extend(data)

    docs: list[dict[str, Any]] = []
    seen = set()
    for item in posts:
        title = html.unescape(item.get("title", ""))
        if "supermarket" not in title.lower() or "price" not in title.lower():
            continue
        post = request_json(POST_URL.format(post_id=item["id"]))
        links = extract_download_links(post["content"]["rendered"])
        for link in links:
            key = (item["id"], link)
            if key in seen:
                continue
            seen.add(key)
            docs.append({"post_id": item["id"], "title": title, "post_url": item["url"], "download_url": link})
    return docs


def month_from_doc(doc: dict[str, Any]) -> datetime | None:
    text = f"{doc['title']} {doc['download_url']}"
    months = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12, "decmber": 12,
    }
    year_match = re.search(r"(20\d{2})", text)
    if not year_match:
        return None
    year = int(year_match.group(1))
    lower = text.lower()
    for name, month in months.items():
        if re.search(rf"\b{name}\b", lower):
            return datetime(year, month, 1, tzinfo=timezone.utc)
    return None


def download(doc: dict[str, Any]) -> Path | None:
    url = doc["download_url"]
    suffix = Path(urllib.parse.urlparse(url).path).suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        return None
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{doc['post_id']}_{Path(urllib.parse.urlparse(url).path).name}"
    path = RAW_DIR / filename
    if not path.exists():
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=120) as response:
            path.write_bytes(response.read())
    return path


def looks_like_label(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    return bool(re.search(r"[A-Za-z]", text))


def parse_price(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip()
    if not text or text in {"-", "N/A", "n/a"}:
        return None
    match = re.search(r"\d+(?:[.,]\d+)?", text.replace("$", ""))
    if not match:
        return None
    return Decimal(match.group(0).replace(",", ".")).quantize(Decimal("0.01"))


def derive_retailer_name(store_label: str) -> str | None:
    norm = normalize(store_label)
    for alias, canonical in RETAILER_ALIASES.items():
        if alias in norm:
            return canonical
    return None


def match_store(store_label: str, stores: list[Any]) -> tuple[Any | None, float]:
    nlabel = normalize(store_label)
    best_store = None
    best_score = 0.0
    for store in stores:
        candidates = [normalize(store.name), normalize(f"{store.retailer.name} {store.name}")]
        score = max(SequenceMatcher(None, nlabel, candidate).ratio() for candidate in candidates)
        if score > best_score:
            best_score = score
            best_store = store
    if best_score >= 0.90:
        return best_store, best_score
    return None, best_score


def get_or_create_retailer(session, name: str | None, retailer_cache: dict[str, Retailer]) -> Retailer | None:
    if not name:
        return None
    if name in retailer_cache:
        return retailer_cache[name]
    retailer = session.scalar(select(Retailer).where(Retailer.name == name))
    if not retailer:
        retailer = Retailer(name=name, integration_type="tradeind", loyalty_program_supported=False)
        session.add(retailer)
        session.flush()
    retailer_cache[name] = retailer
    return retailer


def get_or_create_product(
    session,
    item: str,
    brand: str | None,
    size: str | None,
    family_cache: dict[str, ProductFamily],
    product_cache: dict[tuple[str, str | None], CanonicalProduct],
) -> CanonicalProduct:
    identity = build_product_identity(item, brand=brand, size=size)
    family_name = clean_product_name(item)
    canonical_name = identity.clean_name
    product_key = (identity.normalized_name, identity.normalized_brand)
    if product_key in product_cache:
        return product_cache[product_key]

    family_identity = build_product_identity(item, brand=brand)
    family = family_cache.get(family_name)
    if not family:
        family = session.scalar(select(ProductFamily).where(ProductFamily.name == family_name))
    if not family:
        parsed = parse_package(f"{item} {size or ''}")
        family = ProductFamily(
            name=family_name,
            normalized_name=family_identity.normalized_name,
            selection_key=family_identity.selection_key,
            category="Groceries",
            subcategory=None,
            common_aliases=[],
            default_unit=parsed.normalized_size_unit,
        )
        session.add(family)
        session.flush()
    else:
        family.normalized_name = family.normalized_name or family_identity.normalized_name
        family.selection_key = family.selection_key or family_identity.selection_key
    family_cache[family_name] = family

    product = session.scalar(
        select(CanonicalProduct).where(
            CanonicalProduct.normalized_name == identity.normalized_name,
            CanonicalProduct.normalized_brand == identity.normalized_brand,
        )
    )
    if product:
        if product.size_value is None and identity.parsed_size_value is not None:
            product.size_value = identity.parsed_size_value
            product.size_unit = identity.parsed_size_unit
        if product.package_quantity is None and identity.parsed_package_quantity is not None:
            product.package_quantity = int(identity.parsed_package_quantity)
        product_cache[product_key] = product
        return product

    product = CanonicalProduct(
        product_family_id=family.id,
        canonical_name=canonical_name,
        normalized_name=identity.normalized_name,
        selection_key=identity.selection_key,
        brand=brand,
        normalized_brand=identity.normalized_brand,
        is_store_brand=False,
        owning_retailer_id=None,
        barcode=None,
        category="Groceries",
        subcategory=None,
        size_value=identity.parsed_size_value,
        size_unit=identity.parsed_size_unit,
        package_quantity=int(identity.parsed_package_quantity) if identity.parsed_package_quantity is not None else None,
        tags=[],
        requirements_supported=[],
        is_perishable=False,
    )
    session.add(product)
    session.flush()
    product_cache[product_key] = product
    return product


def iter_workbook_rows(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        region = ws.title.strip()
        if region.lower() not in REGION_SHEETS:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue
        header_index = None
        for idx, candidate in enumerate(rows[:12]):
            normalized_cells = [normalize(cell) for cell in candidate[:6]]
            if "no" in normalized_cells and "items" in normalized_cells:
                header_index = idx
                break
        if header_index is None:
            continue

        area_row = rows[header_index - 1] if header_index > 0 else ()
        header = rows[header_index]
        store_columns = []
        current_area = None
        for index, heading in enumerate(header[4:], start=4):
            area_value = area_row[index] if index < len(area_row) else None
            if area_value:
                current_area = str(area_value).strip()
            if heading and looks_like_label(heading):
                store_columns.append((index, str(heading).strip(), current_area))

        for row in rows[header_index + 1:]:
            item = row[1] if len(row) > 1 else None
            if not item or not looks_like_label(item) or str(item).strip().lower() in {"items", "item"}:
                continue
            brand = row[2] if len(row) > 2 else None
            size = row[3] if len(row) > 3 else None
            for index, store_label, area in store_columns:
                price = parse_price(row[index] if index < len(row) else None)
                if price is None:
                    continue
                yield {
                    "region": region,
                    "area": area,
                    "store_label": store_label,
                    "item": str(item).strip(),
                    "brand": str(brand).strip() if brand else None,
                    "size": str(size).strip() if size else None,
                    "price": price,
                }


def import_docs() -> None:
    docs = discover_price_files()
    xlsx_docs = [doc for doc in docs if doc["download_url"].lower().split("?")[0].endswith((".xlsx", ".xls"))]
    downloaded = []
    for doc in xlsx_docs:
        path = download(doc)
        observed_at = month_from_doc(doc)
        if path and observed_at:
            downloaded.append((doc, path, observed_at))

    inserted = 0
    skipped_existing = 0
    matched_store = 0
    retailer_only = 0
    unmatched = 0
    parse_failures = 0
    skipped_existing_documents = 0

    with SessionLocal() as session:
        stores = session.scalars(select(Store).where(Store.is_active.is_(True))).all()
        retailer_cache: dict[str, Retailer] = {}
        family_cache: dict[str, ProductFamily] = {}
        product_cache: dict[tuple[str, str | None], CanonicalProduct] = {}
        seen_observations: set[tuple[str, datetime, str, str | None, str | None, str]] = set()
        store_match_cache: dict[str, tuple[Store | None, float]] = {}
        for doc, path, observed_at in downloaded:
            already_imported = session.scalar(
                select(ProductPriceObservation.id).where(
                    ProductPriceObservation.source == "tradeind_xlsx",
                    ProductPriceObservation.source_url == doc["download_url"],
                    ProductPriceObservation.observed_at == observed_at,
                ).limit(1)
            )
            if already_imported:
                skipped_existing_documents += 1
                continue

            try:
                rows = list(iter_workbook_rows(path))
            except Exception:
                parse_failures += 1
                continue

            existing_observations = {
                (str(product_id), existing_observed_at, source_url, raw_region, raw_area, raw_store_name)
                for product_id, existing_observed_at, source_url, raw_region, raw_area, raw_store_name in session.execute(
                    select(
                        ProductPriceObservation.canonical_product_id,
                        ProductPriceObservation.observed_at,
                        ProductPriceObservation.source_url,
                        ProductPriceObservation.raw_region,
                        ProductPriceObservation.raw_area,
                        ProductPriceObservation.raw_store_name,
                    ).where(
                        ProductPriceObservation.source == "tradeind_xlsx",
                        ProductPriceObservation.source_url == doc["download_url"],
                        ProductPriceObservation.observed_at == observed_at,
                    )
                )
            }

            for row in rows:
                product = get_or_create_product(session, row["item"], row["brand"], row["size"], family_cache, product_cache)
                store_cache_key = row["store_label"]
                if store_cache_key in store_match_cache:
                    store, confidence = store_match_cache[store_cache_key]
                else:
                    store, confidence = match_store(row["store_label"], stores)
                    store_match_cache[store_cache_key] = (store, confidence)
                retailer_name = derive_retailer_name(row["store_label"])
                retailer = store.retailer if store else get_or_create_retailer(session, retailer_name, retailer_cache)
                if store:
                    matched_store += 1
                elif retailer:
                    retailer_only += 1
                else:
                    unmatched += 1

                observation_key = (
                    str(product.id),
                    observed_at,
                    doc["download_url"],
                    row["region"],
                    row["area"],
                    row["store_label"],
                )
                if observation_key in seen_observations or observation_key in existing_observations:
                    skipped_existing += 1
                    continue
                seen_observations.add(observation_key)

                parsed = parse_package(f"{row['item']} {row['size'] or ''}", price=row["price"])
                session.add(
                    ProductPriceObservation(
                        canonical_product_id=product.id,
                        retailer_id=retailer.id if retailer else None,
                        store_id=store.id if store else None,
                        region_code="TT",
                        price=row["price"],
                        is_on_sale=False,
                        regular_price=None,
                        promotional_tags=[],
                        currency="TTD",
                        price_per_unit=parsed.computed_price_per_unit,
                        observed_at=observed_at,
                        source="tradeind_xlsx",
                        source_url=doc["download_url"],
                        raw_region=row["region"],
                        raw_area=row["area"],
                        raw_store_name=row["store_label"],
                        raw_item_name=row["item"],
                        match_confidence=confidence,
                        raw_payload={
                            **{key: (str(value) if isinstance(value, Decimal) else value) for key, value in row.items()},
                            "post_title": doc["title"],
                            "post_url": doc["post_url"],
                            "file": str(path.name),
                        },
                    )
                )
                inserted += 1
            session.commit()

    print(
        json.dumps(
            {
                "discovered_documents": len(docs),
                "xlsx_documents": len(xlsx_docs),
                "downloaded_with_month": len(downloaded),
                "documents_skipped_existing": skipped_existing_documents,
                "observations_inserted": inserted,
                "observations_skipped_existing": skipped_existing,
                "matched_store_observations": matched_store,
                "retailer_only_observations": retailer_only,
                "unmatched_observations": unmatched,
                "parse_failures": parse_failures,
                "raw_dir": str(RAW_DIR),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    import_docs()
